"""Prepare the Project 2 ontology-generation dataset and resource manifests.

The conversion is intentionally conservative:

* only explicit StoryID values are used to group competency questions;
* semicolon-separated StoryIDs are accepted only when every referenced ID
  exists in the Story sheet;
* gold modules are matched by the dataset README's explicit CQID == filename
  stem rule, case-insensitively;
* missing stories, mappings, prompts, and parse failures are reported rather
  than repaired or inferred.

The script uses only the Python standard library for XLSX/ZIP processing and
RDFLib (already required by the PR #10 baseline) for RDF syntax checks.
"""

from __future__ import annotations

import argparse
import ast
import csv
import hashlib
import json
import os
import posixpath
import re
import shutil
import tempfile
import unicodedata
import zipfile
from collections import Counter, OrderedDict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Sequence
from xml.etree import ElementTree as ET

from rdflib import Graph
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_ARCHIVE_PASSWORD = "28mRFhW6wVnu7Wh"
DEFAULT_APPROVED_METHODS = ("ontogenia", "domain-ontogen", "neon-gpt")

AUTHORITATIVE_PROMPT_REPOSITORIES = {
    "domain-ontogen": {
        "repository": "https://github.com/dersuchendee/Domain-OntoGen",
        "local_path": "external_resources/Domain-OntoGen",
        "branch": "main",
        "commit_sha": "894441e367acdbbd1ea662b6f1a6919d13533051",
        "retrieval_date": "2026-07-15",
        "licence": "Not declared in the repository (no licence file or README statement)",
    },
    "neon-gpt": {
        "repository": "https://github.com/andreamust/NEON-GPT",
        "local_path": "external_resources/NEON-GPT",
        "branch": "main",
        "commit_sha": "bce7a6a805faa23dc169f691afb5aaaacad3d99d",
        "retrieval_date": "2026-07-15",
        "licence": "MIT (repository LICENSE)",
    },
}

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

PROMPT_SPECS = {
    "ontogenia": {
        "expected_adapter_path": (
            "datasets/ontology_generation/raw/ontogenia/"
            "memoryless_cqbycq_prompt.txt"
        ),
        "canonical_path": (
            "datasets/ontology_generation/prompts/ontogenia/P0_original.txt"
        ),
        "required_placeholders": ["{story}", "{CQ}", "{rdf}"],
        "source_kind": "memoryless_readme",
    },
    "domain-ontogen": {
        "expected_adapter_path": (
            "datasets/ontology_generation/raw/domain-ontogen/prompt.txt"
        ),
        "canonical_path": (
            "datasets/ontology_generation/prompts/domain-ontogen/P0_original.txt"
        ),
        "required_placeholders": ["{OS}", "{CQ}"],
        "source_kind": "domain_readme",
        "source_relative_path": "external_resources/Domain-OntoGen/README.md",
        "extraction_boundaries": (
            "README.md section 'Prompt used for ontology generation', "
            "Python fenced string literal"
        ),
    },
    "neon-gpt": {
        "expected_adapter_path": (
            "datasets/ontology_generation/raw/neon-gpt/day1_gpt_prompt_list.txt"
        ),
        "canonical_path": (
            "datasets/ontology_generation/prompts/neon-gpt/P0_original.txt"
        ),
        "required_placeholders": [],
        "source_kind": "byte_copy",
        "source_relative_path": (
            "external_resources/NEON-GPT/gpt_wine_ont_day1/"
            "day1_gpt_prompt_list.txt"
        ),
    },
}


@dataclass(frozen=True)
class WorkbookSheet:
    name: str
    max_row: int
    max_col: int
    rows: dict[int, list[Any]]


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json_hash(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return sha256_bytes(payload)


def atomic_write_bytes(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def atomic_write_text(path: Path, text: str) -> None:
    atomic_write_bytes(path, text.encode("utf-8"))


def write_json(path: Path, value: Any) -> None:
    atomic_write_text(
        path, json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )


def write_csv(path: Path, fieldnames: Sequence[str], rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def column_index(reference: str) -> int:
    letters = re.match(r"[A-Za-z]+", reference)
    if not letters:
        raise ValueError(f"Invalid cell reference: {reference}")
    value = 0
    for char in letters.group(0).upper():
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(
            node.text or "" for node in cell.findall(f".//{{{MAIN_NS}}}t")
        )
    value_node = cell.find(f"{{{MAIN_NS}}}v")
    if value_node is None:
        return None
    raw = value_node.text or ""
    if cell_type == "s":
        return shared_strings[int(raw)]
    if cell_type in {"str", "e"}:
        return raw
    if cell_type == "b":
        return raw == "1"
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def read_xlsx(path: Path) -> list[WorkbookSheet]:
    """Read cell values and physical row positions from a simple XLSX file."""

    with zipfile.ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(f"{{{MAIN_NS}}}si"):
                shared_strings.append(
                    "".join(node.text or "" for node in item.findall(f".//{{{MAIN_NS}}}t"))
                )

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        targets = {
            relation.attrib["Id"]: relation.attrib["Target"]
            for relation in relationships.findall(f"{{{PKG_REL_NS}}}Relationship")
        }

        sheets: list[WorkbookSheet] = []
        for sheet in workbook.findall(f".//{{{MAIN_NS}}}sheet"):
            name = sheet.attrib["name"]
            relationship_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
            target = targets[relationship_id].replace("\\", "/")
            if target.startswith("/"):
                archive_path = target.lstrip("/")
            else:
                archive_path = posixpath.normpath(posixpath.join("xl", target))
            sheet_xml = ET.fromstring(archive.read(archive_path))

            # Do not trust worksheet dimensions: formatting commonly expands
            # them to Excel's final row.  Explicit table ranges are reliable
            # source boundaries and also preserve trailing blank table rows.
            table_max_row = 1
            table_max_col = 1
            sheet_rels_path = posixpath.join(
                posixpath.dirname(archive_path),
                "_rels",
                posixpath.basename(archive_path) + ".rels",
            )
            if sheet_rels_path in archive.namelist():
                sheet_relationships = ET.fromstring(archive.read(sheet_rels_path))
                sheet_targets = {
                    relation.attrib["Id"]: relation.attrib["Target"]
                    for relation in sheet_relationships.findall(
                        f"{{{PKG_REL_NS}}}Relationship"
                    )
                }
                for table_part in sheet_xml.findall(f".//{{{MAIN_NS}}}tablePart"):
                    table_id = table_part.attrib.get(f"{{{DOC_REL_NS}}}id")
                    if not table_id or table_id not in sheet_targets:
                        continue
                    table_path = posixpath.normpath(
                        posixpath.join(
                            posixpath.dirname(archive_path), sheet_targets[table_id]
                        )
                    )
                    table_root = ET.fromstring(archive.read(table_path))
                    table_last_ref = table_root.attrib.get("ref", "A1").split(":")[-1]
                    row_match = re.search(r"(\d+)$", table_last_ref)
                    if row_match:
                        table_max_row = max(table_max_row, int(row_match.group(1)))
                    table_max_col = max(table_max_col, column_index(table_last_ref))

            max_row = 1
            max_col = 1

            rows: dict[int, list[Any]] = {}
            for row in sheet_xml.findall(f".//{{{MAIN_NS}}}row"):
                row_number = int(row.attrib["r"])
                cells: dict[int, Any] = {}
                for cell in row.findall(f"{{{MAIN_NS}}}c"):
                    col = column_index(cell.attrib["r"])
                    cells[col] = _xlsx_cell_value(cell, shared_strings)
                    max_col = max(max_col, col)
                rows[row_number] = [cells.get(index) for index in range(1, max_col + 1)]
                max_row = max(max_row, row_number)
            meaningful_rows = [
                row_number
                for row_number, values in rows.items()
                if any(cell_text(value) for value in values)
            ]
            # XLSX dimensions frequently span entire formatted columns.  The
            # meaningful data boundary is the final row containing a value;
            # blank rows inside that boundary remain addressable by source row.
            max_row = max(max(meaningful_rows, default=1), table_max_row)
            max_col = max(max_col, table_max_col)
            sheets.append(
                WorkbookSheet(
                    name=name,
                    max_row=max_row,
                    max_col=max_col,
                    rows=rows,
                )
            )
        return sheets


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_dict(sheet: WorkbookSheet, row_number: int) -> dict[str, Any]:
    headers = [cell_text(value) for value in sheet.rows.get(1, [])]
    values = sheet.rows.get(row_number, [])
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(headers)
        if header
    }


def sanitize_filename(value: str, fallback: str = "resource") -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return (normalized or fallback)[:120]


def portable_path(path: Path, repo_root: Path) -> str:
    """Return a stable repository-relative path whenever possible."""

    resolved = path.resolve()
    try:
        return resolved.relative_to(repo_root.resolve()).as_posix()
    except ValueError:
        return str(resolved).replace("\\", "/")


def dataset_id_for_story(story_id: str) -> str:
    slug = sanitize_filename(story_id, "story")[:64]
    suffix = hashlib.sha256(story_id.encode("utf-8")).hexdigest()[:10]
    return f"project2_{slug}_{suffix}"


def rdf_format_for_path(path: Path) -> str:
    return {
        ".ttl": "turtle",
        ".rdf": "xml",
        ".owl": "xml",
        ".jsonld": "json-ld",
        ".json-ld": "json-ld",
        ".nt": "nt",
    }.get(path.suffix.lower(), "turtle")


def parse_rdf(path: Path) -> tuple[bool, str | None, int | None, str | None]:
    """Parse RDF with content-aware fallback for mislabeled source files."""

    leading = path.read_bytes()[:512].lstrip().lower()
    extension_format = rdf_format_for_path(path)
    candidates = [extension_format]
    if leading.startswith(b"<?xml") or leading.startswith(b"<rdf:rdf"):
        candidates = ["xml", extension_format]
    elif leading.startswith((b"{", b"[")):
        candidates = ["json-ld", extension_format]
    candidates.extend(["turtle", "xml", "json-ld", "nt"])

    errors: list[str] = []
    for rdf_format in dict.fromkeys(candidates):
        try:
            graph = Graph()
            graph.parse(path, format=rdf_format)
            return True, None, len(graph), rdf_format
        except Exception as exc:  # RDFLib exposes parser-specific errors.
            errors.append(f"{rdf_format}: {exc}")
    return False, " | ".join(errors), None, None


def safe_archive_member(member: str) -> PurePosixPath:
    candidate = PurePosixPath(member)
    if candidate.is_absolute() or ".." in candidate.parts:
        raise ValueError(f"Unsafe archive member path: {member}")
    return candidate


def extract_gold_archives(
    dataset_source: Path,
    gold_output: Path,
    password: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    password_bytes = password.encode("utf-8")
    for archive_path in sorted(dataset_source.glob("*.zip")):
        archive_dir = sanitize_filename(archive_path.stem)
        with zipfile.ZipFile(archive_path) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                try:
                    member = safe_archive_member(info.filename)
                except ValueError as exc:
                    errors.append(
                        {
                            "category": "unsafe_archive_member",
                            "source": str(archive_path),
                            "identifier": info.filename,
                            "detail": str(exc),
                        }
                    )
                    continue
                if member.suffix.lower() not in {".ttl", ".rdf", ".owl"}:
                    errors.append(
                        {
                            "category": "unsupported_gold_file",
                            "source": str(archive_path),
                            "identifier": info.filename,
                            "detail": "Only .ttl, .rdf, and .owl are extracted.",
                        }
                    )
                    continue
                try:
                    content = archive.read(info, pwd=password_bytes)
                except Exception as exc:
                    errors.append(
                        {
                            "category": "archive_extraction_error",
                            "source": str(archive_path),
                            "identifier": info.filename,
                            "detail": str(exc),
                        }
                    )
                    continue
                output_name = sanitize_filename(member.name, "gold.ttl")
                output_path = gold_output / archive_dir / output_name
                atomic_write_bytes(output_path, content)
                parse_success, parse_error, triple_count, parse_format = parse_rdf(
                    output_path
                )
                records.append(
                    {
                        "archive": archive_path.name,
                        "archive_member": info.filename,
                        "source_sha256": sha256_bytes(content),
                        "filename": output_name,
                        "stem": Path(output_name).stem,
                        "output_path": output_path,
                        "parse_success": parse_success,
                        "parse_error": parse_error,
                        "parse_format": parse_format,
                        "triple_count": triple_count,
                    }
                )
                if not parse_success:
                    errors.append(
                        {
                            "category": "unparseable_gold_file",
                            "source": str(archive_path),
                            "identifier": info.filename,
                            "detail": parse_error or "Unknown RDF parse error",
                        }
                    )
    return records, errors


def extract_memoryless_prompt(markdown_path: Path) -> str:
    text = markdown_path.read_text(encoding="utf-8")
    section_match = re.search(
        r"(?s)^##\s+Memoryless\s+CQbyCQ\s*(.*?)(?=^##\s+Ontogenia\b)",
        text,
        re.MULTILINE,
    )
    if not section_match:
        raise ValueError(f"Memoryless CQbyCQ section not found in {markdown_path}")
    code_match = re.search(r"(?s)```python\s*(.*?)\s*```", section_match.group(1))
    if not code_match:
        raise ValueError(f"Memoryless prompt code block not found in {markdown_path}")
    expression = code_match.group(1).strip()
    try:
        prompt = ast.literal_eval(expression)
    except Exception as exc:
        raise ValueError(f"Memoryless prompt is not one string literal: {exc}") from exc
    if not isinstance(prompt, str):
        raise ValueError("Memoryless prompt literal did not evaluate to text")
    return prompt.rstrip() + "\n"


def extract_domain_ontogen_prompt(markdown_path: Path) -> str:
    """Evaluate the exact published Python string under the named README section."""

    text = markdown_path.read_text(encoding="utf-8")
    section_match = re.search(
        r"(?s)^##\s+Prompt used for ontology generation:\s*(.*?)(?=^##\s+|\Z)",
        text,
        re.MULTILINE,
    )
    if not section_match:
        raise ValueError(
            f"Prompt used for ontology generation section not found in {markdown_path}"
        )
    code_match = re.search(r"(?s)```python\s*(.*?)\s*```", section_match.group(1))
    if not code_match:
        raise ValueError(f"Published Python prompt block not found in {markdown_path}")
    expression = code_match.group(1).strip()
    try:
        prompt = ast.literal_eval(expression)
    except Exception as exc:
        raise ValueError(f"Published prompt is not one Python string literal: {exc}") from exc
    if not isinstance(prompt, str) or not prompt:
        raise ValueError("Published Domain-OntoGen prompt is empty or not text")
    return prompt


def prepare_prompts(
    repo_root: Path,
    staging_root: Path,
    approved_methods: Sequence[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    prompt_records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    source_readme = (
        repo_root
        / "external_resources"
        / "Onto-Generation"
        / "PromptingTechniques"
        / "README.md"
    )
    for method in approved_methods:
        spec = PROMPT_SPECS.get(method)
        if not spec:
            errors.append(
                {
                    "category": "unknown_approved_method",
                    "source": "config/course_methods.yaml",
                    "identifier": method,
                    "detail": "No prompt preparation specification exists.",
                }
            )
            continue

        prompt_bytes: bytes | None = None
        source_path: Path | None = None
        provenance: dict[str, Any] = {}
        if spec["source_kind"] == "memoryless_readme":
            source_path = source_readme
            try:
                prompt_bytes = extract_memoryless_prompt(source_readme).encode("utf-8")
                provenance = {
                    "classification": "published_exact",
                    "extraction_boundaries": (
                        "README.md section 'Memoryless CQbyCQ', Python fenced string literal"
                    ),
                    "transformations": {
                        "markdown_fence_removed": True,
                        "python_string_delimiters_removed": True,
                        "python_escape_sequences_decoded": True,
                        "semantic_modification": False,
                    },
                }
            except Exception as exc:
                errors.append(
                    {
                        "category": "prompt_extraction_error",
                        "source": str(source_readme),
                        "identifier": method,
                        "detail": str(exc),
                    }
                )
        elif spec["source_kind"] == "domain_readme":
            source_path = repo_root / spec["source_relative_path"]
            try:
                prompt_bytes = extract_domain_ontogen_prompt(source_path).encode("utf-8")
                provenance = {
                    **AUTHORITATIVE_PROMPT_REPOSITORIES[method],
                    "classification": "published_exact",
                    "source_file": spec["source_relative_path"],
                    "source_readme_sha256": sha256_file(source_path),
                    "extraction_boundaries": spec["extraction_boundaries"],
                    "only_markdown_indentation_or_fence_removed": False,
                    "transformations": {
                        "markdown_fence_removed": True,
                        "python_string_delimiters_removed": True,
                        "python_escape_sequences_decoded": True,
                        "semantic_modification": False,
                    },
                }
            except Exception as exc:
                errors.append(
                    {
                        "category": "prompt_extraction_error",
                        "source": str(source_path),
                        "identifier": method,
                        "detail": str(exc),
                    }
                )
        elif spec["source_kind"] == "byte_copy":
            source_path = repo_root / spec["source_relative_path"]
            try:
                prompt_bytes = source_path.read_bytes()
                provenance = {
                    **AUTHORITATIVE_PROMPT_REPOSITORIES[method],
                    "classification": "published_exact",
                    "source_file": spec["source_relative_path"],
                    "copy_mode": "byte-for-byte",
                    "source_sha256": sha256_bytes(prompt_bytes),
                }
            except Exception as exc:
                errors.append(
                    {
                        "category": "prompt_copy_error",
                        "source": str(source_path),
                        "identifier": method,
                        "detail": str(exc),
                    }
                )
        else:
            errors.append(
                {
                    "category": "unsupported_prompt_source_kind",
                    "source": "scripts/prepare_ontology_dataset.py",
                    "identifier": method,
                    "detail": spec["source_kind"],
                }
            )

        if prompt_bytes is None:
            expected_source = spec.get("source_relative_path", "authoritative README")
            errors.append(
                {
                    "category": "missing_prompt",
                    "source": "authoritative repositories",
                    "identifier": method,
                    "detail": (
                        f"Expected {spec['expected_adapter_path']}; source: "
                        f"{expected_source}"
                    ),
                }
            )
            prompt_records.append(
                {
                    "method": method,
                    "status": "missing",
                    "expected_adapter_path": spec["expected_adapter_path"],
                    "canonical_path": spec["canonical_path"],
                    "source_path": None,
                    "sha256": None,
                    "required_placeholders": spec["required_placeholders"],
                    "missing_placeholders": spec["required_placeholders"],
                }
            )
            continue

        prompt_text = prompt_bytes.decode("utf-8")
        missing_placeholders = [
            placeholder
            for placeholder in spec["required_placeholders"]
            if placeholder not in prompt_text
        ]
        if missing_placeholders:
            errors.append(
                {
                    "category": "prompt_placeholder_error",
                    "source": str(source_path),
                    "identifier": method,
                    "detail": f"Missing placeholders: {missing_placeholders}",
                }
            )
        canonical_relative = Path(spec["canonical_path"]).relative_to(
            "datasets/ontology_generation"
        )
        adapter_relative = Path(spec["expected_adapter_path"]).relative_to(
            "datasets/ontology_generation"
        )
        canonical_output = staging_root / canonical_relative
        adapter_output = staging_root / adapter_relative
        atomic_write_bytes(canonical_output, prompt_bytes)
        atomic_write_bytes(adapter_output, prompt_bytes)
        copied_sha256 = sha256_file(adapter_output)
        prompt_records.append(
            {
                "method": method,
                "status": "ready" if not missing_placeholders else "invalid",
                "expected_adapter_path": spec["expected_adapter_path"],
                "canonical_path": spec["canonical_path"],
                "source_path": str(source_path),
                "source_sha256": sha256_file(source_path),
                "sha256": sha256_file(canonical_output),
                "copied_sha256": copied_sha256,
                "source_and_copy_hashes_match": (
                    sha256_file(source_path) == copied_sha256
                    if spec["source_kind"] == "byte_copy"
                    else None
                ),
                "required_placeholders": spec["required_placeholders"],
                "missing_placeholders": missing_placeholders,
                "placeholder_counts": {
                    placeholder: prompt_text.count(placeholder)
                    for placeholder in spec["required_placeholders"]
                },
                "provenance": provenance,
            }
        )
    return prompt_records, errors


def convert_patterns_to_odps(
    patterns_csv: Path,
    output_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Preserve each Pattern_owl cell as a separately loadable `.owl` file."""

    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    used_names: Counter[str] = Counter()
    with patterns_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"Name", "Pattern_owl"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise ValueError(
                f"patterns CSV requires columns {sorted(required)}; got {reader.fieldnames}"
            )
        for source_row, row in enumerate(reader, start=2):
            original_name = cell_text(row.get("Name"))
            content = cell_text(row.get("Pattern_owl"))
            base = sanitize_filename(original_name, f"pattern_{source_row}")
            used_names[base.casefold()] += 1
            occurrence = used_names[base.casefold()]
            filename = f"{base}.owl" if occurrence == 1 else f"{base}_{occurrence}.owl"
            output_path = output_dir / filename
            if not content:
                parse_success, parse_error, triple_count = (
                    False,
                    "Empty Pattern_owl value",
                    None,
                )
                atomic_write_text(output_path, "")
            else:
                atomic_write_text(output_path, content.rstrip() + "\n")
            parse_success, parse_error, triple_count, parse_format = parse_rdf(
                output_path
            )
            record = {
                "source_row": source_row,
                "original_name": original_name,
                "sanitized_filename": filename,
                "output_path": str(output_path),
                "sha256": sha256_file(output_path),
                "parse_success": parse_success,
                    "parse_error": parse_error,
                    "parse_format": parse_format,
                    "triple_count": triple_count,
            }
            records.append(record)
            if not parse_success:
                errors.append(
                    {
                        "category": "unparseable_odp_file",
                        "source": str(patterns_csv),
                        "identifier": original_name or str(source_row),
                        "detail": parse_error or "Unknown RDF parse error",
                    }
                )
    return records, errors


def sheet_audit(sheet: WorkbookSheet) -> dict[str, Any]:
    headers = [cell_text(value) for value in sheet.rows.get(1, [])]
    missing = {header: 0 for header in headers if header}
    for row_number in range(2, sheet.max_row + 1):
        row = row_dict(sheet, row_number)
        for header in missing:
            if not cell_text(row.get(header)):
                missing[header] += 1
    return {
        "name": sheet.name,
        "max_row": sheet.max_row,
        "data_row_count": max(0, sheet.max_row - 1),
        "column_names": headers,
        "missing_values_by_column": missing,
    }


def normalize_identifier(value: Any) -> tuple[str, bool]:
    """Apply only the Phase 2B-approved representation normalizations."""

    if value is None:
        return "", False
    if isinstance(value, float) and value.is_integer():
        return str(int(value)), True
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value), True
    raw = str(value)
    normalized = unicodedata.normalize("NFC", raw.strip()).replace("\r\n", "\n").replace(
        "\r", "\n"
    )
    return normalized, normalized != raw


def normalize_source_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value).strip()).replace(
        "\r\n", "\n"
    ).replace("\r", "\n")


def compress_integer_ranges(values: Sequence[int]) -> list[str]:
    values = sorted(set(values))
    if not values:
        return []
    ranges: list[str] = []
    start = previous = values[0]
    for value in values[1:]:
        if value == previous + 1:
            previous = value
            continue
        ranges.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = value
    ranges.append(str(start) if start == previous else f"{start}-{previous}")
    return ranges


def inspect_workbook_with_openpyxl(
    workbook_path: Path,
    cq_max_row: int,
    story_max_row: int,
) -> dict[str, Any]:
    """Inspect Excel structure without using formatting as mapping evidence."""

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    result: dict[str, Any] = {
        "library": "openpyxl",
        "openpyxl_inspection_performed": True,
        "sheets": {},
        "cq_story_cells": {},
        "story_id_cells": {},
    }
    boundaries = {"CQs": cq_max_row, "Story": story_max_row}
    relevant_columns = {"CQs": 4, "Story": 2}
    for sheet_name, boundary in boundaries.items():
        worksheet = workbook[sheet_name]
        merged_ranges = sorted(
            (str(item) for item in worksheet.merged_cells.ranges),
            key=str.casefold,
        )
        story_column_merges = sorted(
            (
                str(item)
                for item in worksheet.merged_cells.ranges
                if item.min_col <= 1 <= item.max_col
                and item.max_row >= 2
                and item.min_row <= boundary
            ),
            key=str.casefold,
        )
        hidden_rows = [
            row
            for row in range(2, boundary + 1)
            if bool(worksheet.row_dimensions[row].hidden)
        ]
        hidden_columns = [
            get_column_letter(column)
            for column in range(1, relevant_columns[sheet_name] + 1)
            if bool(worksheet.column_dimensions[get_column_letter(column)].hidden)
        ]
        formulas = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=boundary,
            min_col=1,
            max_col=relevant_columns[sheet_name],
        ):
            for cell in row:
                if cell.data_type == "f":
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
        result["sheets"][sheet_name] = {
            "source_boundary_row": boundary,
            "merged_ranges": merged_ranges,
            "story_id_column_merged_ranges": story_column_merges,
            "hidden_rows": hidden_rows,
            "hidden_row_ranges": compress_integer_ranges(hidden_rows),
            "hidden_columns": hidden_columns,
            "formula_cells": formulas,
        }

    cq_sheet = workbook["CQs"]
    for row_number in range(2, cq_max_row + 1):
        cell = cq_sheet.cell(row_number, 1)
        raw_value = cell.value
        merged_source_range = None
        merged_source_cell = None
        for merged_range in cq_sheet.merged_cells.ranges:
            if merged_range.min_col <= 1 <= merged_range.max_col and (
                merged_range.min_row <= row_number <= merged_range.max_row
            ):
                merged_source_range = str(merged_range)
                merged_source_cell = cq_sheet.cell(
                    merged_range.min_row, merged_range.min_col
                ).coordinate
                if raw_value in (None, ""):
                    raw_value = cq_sheet.cell(
                        merged_range.min_row, merged_range.min_col
                    ).value
                break
        normalized, normalization_applied = normalize_identifier(raw_value)
        result["cq_story_cells"][row_number] = {
            "raw_value": cell.value,
            "raw_value_type": type(cell.value).__name__,
            "normalized_value": normalized,
            "normalization_applied": normalization_applied,
            "merged_source_range": merged_source_range,
            "merged_source_cell": merged_source_cell,
            "merged_cell_propagation_applied": (
                merged_source_range is not None
                and cell.value in (None, "")
                and raw_value not in (None, "")
            ),
            "hidden_row": bool(cq_sheet.row_dimensions[row_number].hidden),
        }

    story_sheet = workbook["Story"]
    for row_number in range(2, story_max_row + 1):
        value = story_sheet.cell(row_number, 1).value
        normalized, normalization_applied = normalize_identifier(value)
        result["story_id_cells"][row_number] = {
            "raw_value": value,
            "raw_value_type": type(value).__name__,
            "normalized_value": normalized,
            "normalization_applied": normalization_applied,
        }
    return result


def build_dataset_records(
    workbook_path: Path,
    gold_records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    sheets = read_xlsx(workbook_path)
    by_name = {sheet.name: sheet for sheet in sheets}
    if "CQs" not in by_name or "Story" not in by_name:
        raise ValueError(
            f"Workbook must contain CQs and Story sheets; found {sorted(by_name)}"
        )
    cq_sheet = by_name["CQs"]
    story_sheet = by_name["Story"]
    cq_headers = [cell_text(value) for value in cq_sheet.rows.get(1, [])]
    story_headers = [cell_text(value) for value in story_sheet.rows.get(1, [])]
    required_cq = {"StoryID", "CQID", "CQText", "Category of CQ"}
    required_story = {"StoryID", "StoryText"}
    if not required_cq.issubset(cq_headers):
        raise ValueError(f"CQs sheet missing columns: {sorted(required_cq - set(cq_headers))}")
    if not required_story.issubset(story_headers):
        raise ValueError(
            f"Story sheet missing columns: {sorted(required_story - set(story_headers))}"
        )

    stories: OrderedDict[str, dict[str, Any]] = OrderedDict()
    story_id_rows: dict[str, list[int]] = {}
    blank_story_rows: list[int] = []
    for source_row in range(2, story_sheet.max_row + 1):
        source = row_dict(story_sheet, source_row)
        story_id = cell_text(source.get("StoryID"))
        story_text = cell_text(source.get("StoryText"))
        if not story_id and not story_text:
            blank_story_rows.append(source_row)
            continue
        if story_id:
            story_id_rows.setdefault(story_id.casefold(), []).append(source_row)
            stories.setdefault(
                story_id.casefold(),
                {
                    "story_id": story_id,
                    "story_text": story_text,
                    "source_row": source_row,
                },
            )

    duplicate_story_ids = [
        {"story_id": stories[key]["story_id"], "source_rows": rows}
        for key, rows in story_id_rows.items()
        if len(rows) > 1
    ]

    raw_cq_rows: list[dict[str, Any]] = []
    blank_cq_rows: list[int] = []
    for source_row in range(2, cq_sheet.max_row + 1):
        source = row_dict(cq_sheet, source_row)
        record = {
            "source_row": source_row,
            "story_id": cell_text(source.get("StoryID")),
            "cq_id": cell_text(source.get("CQID")),
            "cq_text": cell_text(source.get("CQText")),
            "category": cell_text(source.get("Category of CQ")),
        }
        if not any(record[key] for key in ("story_id", "cq_id", "cq_text", "category")):
            blank_cq_rows.append(source_row)
        raw_cq_rows.append(record)

    cq_counts = Counter(
        row["cq_id"].casefold() for row in raw_cq_rows if row["cq_id"]
    )
    duplicate_cq_ids = [
        {
            "cq_id": next(
                row["cq_id"]
                for row in raw_cq_rows
                if row["cq_id"] and row["cq_id"].casefold() == key
            ),
            "source_rows": [
                row["source_row"]
                for row in raw_cq_rows
                if row["cq_id"] and row["cq_id"].casefold() == key
            ],
        }
        for key, count in sorted(cq_counts.items())
        if count > 1
    ]

    gold_by_stem: dict[str, list[dict[str, Any]]] = {}
    for record in gold_records:
        gold_by_stem.setdefault(record["stem"].casefold(), []).append(record)

    excluded: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []
    all_source_cq_ids = {
        row["cq_id"].casefold() for row in raw_cq_rows if row["cq_id"]
    }
    mapping_rows: list[dict[str, Any]] = []

    for row in raw_cq_rows:
        reasons: list[str] = []
        if row["source_row"] in blank_cq_rows:
            reasons.append("blank_row")
        else:
            if not row["cq_id"]:
                reasons.append("missing_cq_id")
            if not row["cq_text"]:
                reasons.append("empty_cq_text")
            if not row["story_id"]:
                reasons.append("missing_story_id")
            if row["cq_id"] and cq_counts[row["cq_id"].casefold()] > 1:
                reasons.append("duplicate_cq_id")

        story_ids = [
            part.strip() for part in row["story_id"].split(";") if part.strip()
        ]
        missing_references = [
            story_id
            for story_id in story_ids
            if story_id.casefold() not in stories
        ]
        if missing_references:
            reasons.append("unresolved_story_id")

        gold_candidates = gold_by_stem.get(row["cq_id"].casefold(), []) if row["cq_id"] else []
        if len(gold_candidates) == 1:
            gold = gold_candidates[0]
            mapping_status = "mapped"
            gold_path = str(gold["output_path"])
            gold_sha = gold["source_sha256"]
            gold_parse_success = gold["parse_success"]
        elif len(gold_candidates) > 1:
            mapping_status = "ambiguous"
            gold_path = "|".join(str(candidate["output_path"]) for candidate in gold_candidates)
            gold_sha = "|".join(candidate["source_sha256"] for candidate in gold_candidates)
            gold_parse_success = False
            reasons.append("ambiguous_gold_mapping")
        else:
            mapping_status = "missing"
            gold_path = ""
            gold_sha = ""
            gold_parse_success = ""

        included = not reasons
        mapping_rows.append(
            {
                "source_workbook": workbook_path.name,
                "source_sheet": "CQs",
                "source_row": row["source_row"],
                "story_id": row["story_id"],
                "cq_id": row["cq_id"],
                "included_in_normalized_dataset": str(included).lower(),
                "mapping_rule": "case-insensitive exact CQID-to-filename-stem",
                "mapping_status": mapping_status,
                "gold_path": gold_path,
                "gold_sha256": gold_sha,
                "gold_parse_success": gold_parse_success,
                "exclusion_reasons": "|".join(reasons),
            }
        )

        enriched = {
            **row,
            "story_ids": story_ids,
            "story_records": [stories[story_id.casefold()] for story_id in story_ids if story_id.casefold() in stories],
            "gold_path": gold_path or None,
            "gold_sha256": gold_sha or None,
            "gold_parse_success": gold_parse_success if gold_path else None,
            "mapping_status": mapping_status,
        }
        if reasons:
            excluded_entry = {
                "source_sheet": "CQs",
                "source_row": row["source_row"],
                "story_id": row["story_id"],
                "cq_id": row["cq_id"],
                "reasons": reasons,
            }
            excluded.append(excluded_entry)
            errors.append(
                {
                    "category": "excluded_source_row",
                    "source": f"{workbook_path.name}:CQs",
                    "identifier": str(row["source_row"]),
                    "detail": "|".join(reasons),
                }
            )
        else:
            valid_rows.append(enriched)

    groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for row in valid_rows:
        groups.setdefault(row["story_id"], []).append(row)

    normalized_items: list[dict[str, Any]] = []
    for story_id, rows in groups.items():
        story_records: OrderedDict[str, dict[str, Any]] = OrderedDict()
        for row in rows:
            for story in row["story_records"]:
                story_records.setdefault(story["story_id"], story)
        user_stories = [story["story_text"] for story in story_records.values()]
        if len(story_records) == 1:
            scenario = user_stories[0]
        else:
            scenario = "\n\n".join(
                f"[{story['story_id']}]\n{story['story_text']}"
                for story in story_records.values()
            )
        cq_metadata = [
            {
                "cq_id": row["cq_id"],
                "cq_text": row["cq_text"],
                "category": row["category"] or None,
                "source_row": row["source_row"],
                "gold_module": row["gold_path"],
                "gold_sha256": row["gold_sha256"],
                "gold_parse_success": row["gold_parse_success"],
                "gold_mapping_status": row["mapping_status"],
            }
            for row in rows
        ]
        normalized_items.append(
            {
                "dataset_id": dataset_id_for_story(story_id),
                "scenario_id": story_id,
                "scenario": scenario,
                "competency_questions": [row["cq_text"] for row in rows],
                "user_stories": user_stories,
                "constraints": {"output_format": "ttl"},
                "metadata": {
                    "original_story_id": story_id,
                    "original_story_ids": list(story_records),
                    "original_cq_ids": [row["cq_id"] for row in rows],
                    "source_workbook": workbook_path.name,
                    "source_sheet": "CQs",
                    "source_rows": [row["source_row"] for row in rows],
                    "story_source_rows": [story["source_row"] for story in story_records.values()],
                    "domain_identifier": None,
                    "domain_identifier_source": "not_provided_by_source_workbook",
                    "ontology_identifiers": [row["cq_id"] for row in rows],
                    "cq_records": cq_metadata,
                },
            }
        )

    mapped_gold_paths = {
        Path(row["gold_path"]).resolve()
        for row in mapping_rows
        if row["mapping_status"] == "mapped" and row["gold_path"]
    }
    orphan_modules = [
        {
            "archive": record["archive"],
            "archive_member": record["archive_member"],
            "output_path": str(record["output_path"]),
            "stem": record["stem"],
            "reason": (
                "No CQID with a case-insensitive exact stem match"
                if record["stem"].casefold() not in all_source_cq_ids
                else "Matched CQ exists but mapping was not unique"
            ),
        }
        for record in gold_records
        if record["output_path"].resolve() not in mapped_gold_paths
    ]

    audit = {
        "source_workbooks": [
            {
                "path": str(workbook_path),
                "sha256": sha256_file(workbook_path),
            }
        ],
        "sheets": [sheet_audit(sheet) for sheet in sheets],
        "source_row_count": max(0, cq_sheet.max_row - 1),
        "story_count": len(stories),
        "cq_count": sum(
            1 for row in raw_cq_rows if row["cq_id"] and row["cq_text"]
        ),
        "normalized_item_count": len(normalized_items),
        "normalized_cq_count": sum(
            len(item["competency_questions"]) for item in normalized_items
        ),
        "gold_module_count": len(gold_records),
        "duplicate_ids": {
            "cq_ids": duplicate_cq_ids,
            "story_ids": duplicate_story_ids,
        },
        "missing_ids": {
            "cq_id_source_rows": [row["source_row"] for row in raw_cq_rows if not row["cq_id"]],
            "story_id_source_rows": [row["source_row"] for row in raw_cq_rows if not row["story_id"]],
        },
        "empty_cqs": [row["source_row"] for row in raw_cq_rows if not row["cq_text"]],
        "missing_stories": [
            {
                "source_row": row["source_row"],
                "cq_id": row["cq_id"],
                "story_id": row["story_id"],
            }
            for row in raw_cq_rows
            if row["cq_id"]
            and (
                not row["story_id"]
                or any(
                    part.strip().casefold() not in stories
                    for part in row["story_id"].split(";")
                    if part.strip()
                )
            )
        ],
        "blank_story_rows": blank_story_rows,
        "missing_mappings": [
            {
                "source_row": row["source_row"],
                "story_id": row["story_id"],
                "cq_id": row["cq_id"],
            }
            for row in mapping_rows
            if row["cq_id"] and row["mapping_status"] == "missing"
        ],
        "orphan_modules": orphan_modules,
        "excluded_rows_and_reasons": excluded,
        "domain_identifier_column_present": False,
        "gold_mapping_rule": (
            "Dataset README explicitly states module filename is CQID.ttl; "
            "matching is case-insensitive exact filename stem equality."
        ),
    }
    return normalized_items, mapping_rows, audit, errors


def build_dataset_records_phase2b(
    workbook_path: Path,
    gold_records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    """Build explicit full-generation and exact-gold scopes for Phase 2B."""

    sheets = read_xlsx(workbook_path)
    by_name = {sheet.name: sheet for sheet in sheets}
    if "CQs" not in by_name or "Story" not in by_name:
        raise ValueError(
            f"Workbook must contain CQs and Story sheets; found {sorted(by_name)}"
        )
    cq_sheet = by_name["CQs"]
    story_sheet = by_name["Story"]
    cq_headers = [cell_text(value) for value in cq_sheet.rows.get(1, [])]
    story_headers = [cell_text(value) for value in story_sheet.rows.get(1, [])]
    required_cq = {"StoryID", "CQID", "CQText", "Category of CQ"}
    required_story = {"StoryID", "StoryText"}
    if not required_cq.issubset(cq_headers):
        raise ValueError(f"CQs sheet missing columns: {sorted(required_cq - set(cq_headers))}")
    if not required_story.issubset(story_headers):
        raise ValueError(
            f"Story sheet missing columns: {sorted(required_story - set(story_headers))}"
        )

    structure = inspect_workbook_with_openpyxl(
        workbook_path, cq_sheet.max_row, story_sheet.max_row
    )
    stories_by_key: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    blank_story_rows: list[int] = []
    for source_row in range(2, story_sheet.max_row + 1):
        source = row_dict(story_sheet, source_row)
        cell_metadata = structure["story_id_cells"][source_row]
        story_id = cell_metadata["normalized_value"]
        story_text = normalize_source_text(source.get("StoryText"))
        if not story_id and not story_text:
            blank_story_rows.append(source_row)
            continue
        if story_id:
            stories_by_key.setdefault(story_id.casefold(), []).append(
                {
                    "story_id": story_id,
                    "story_text": story_text,
                    "source_row": source_row,
                    "normalization_applied": cell_metadata["normalization_applied"],
                    "raw_value_type": cell_metadata["raw_value_type"],
                }
            )

    duplicate_story_ids = [
        {
            "story_id": records[0]["story_id"],
            "source_rows": [record["source_row"] for record in records],
        }
        for records in stories_by_key.values()
        if len(records) > 1
    ]

    raw_cq_rows: list[dict[str, Any]] = []
    for source_row in range(2, cq_sheet.max_row + 1):
        source = row_dict(cq_sheet, source_row)
        story_metadata = structure["cq_story_cells"][source_row]
        cq_id, cq_id_normalized = normalize_identifier(source.get("CQID"))
        raw_cq_rows.append(
            {
                "source_row": source_row,
                "raw_story_id": cell_text(story_metadata["raw_value"]),
                "story_id": story_metadata["normalized_value"],
                "story_cell": story_metadata,
                "cq_id": cq_id,
                "cq_id_normalization_applied": cq_id_normalized,
                "cq_text": normalize_source_text(source.get("CQText")),
                "category": normalize_source_text(source.get("Category of CQ")),
            }
        )

    cq_counts = Counter(
        row["cq_id"].casefold() for row in raw_cq_rows if row["cq_id"]
    )
    duplicate_cq_ids = [
        {
            "cq_id": next(
                row["cq_id"]
                for row in raw_cq_rows
                if row["cq_id"] and row["cq_id"].casefold() == key
            ),
            "source_rows": [
                row["source_row"]
                for row in raw_cq_rows
                if row["cq_id"] and row["cq_id"].casefold() == key
            ],
        }
        for key, count in sorted(cq_counts.items())
        if count > 1
    ]

    gold_by_stem: dict[str, list[dict[str, Any]]] = {}
    for record in gold_records:
        normalized_stem, _ = normalize_identifier(record["stem"])
        gold_by_stem.setdefault(normalized_stem.casefold(), []).append(record)

    exclusion_categories = [
        "empty_cq",
        "duplicate_cq",
        "missing_story_id",
        "story_id_not_found",
        "ambiguous_story_mapping",
        "missing_gold_only",
        "invalid_source_row",
        "other",
    ]
    excluded: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    full_generation_rows: list[dict[str, Any]] = []
    mapping_rows: list[dict[str, Any]] = []
    all_source_cq_ids = {
        row["cq_id"].casefold() for row in raw_cq_rows if row["cq_id"]
    }

    for row in raw_cq_rows:
        story_ids = [
            normalize_identifier(part)[0]
            for part in row["story_id"].split(";")
            if normalize_identifier(part)[0]
        ]
        missing_story_ids = [
            story_id
            for story_id in story_ids
            if story_id.casefold() not in stories_by_key
        ]
        ambiguous_story_ids = [
            story_id
            for story_id in story_ids
            if len(stories_by_key.get(story_id.casefold(), [])) > 1
        ]

        if not row["cq_text"]:
            exclusion_reason = "empty_cq"
        elif not row["cq_id"]:
            exclusion_reason = "invalid_source_row"
        elif cq_counts[row["cq_id"].casefold()] > 1:
            exclusion_reason = "duplicate_cq"
        elif not row["story_id"]:
            exclusion_reason = "missing_story_id"
        elif ambiguous_story_ids:
            exclusion_reason = "ambiguous_story_mapping"
        elif missing_story_ids:
            exclusion_reason = "story_id_not_found"
        else:
            exclusion_reason = ""

        if exclusion_reason:
            mapping_method = "unmapped"
        elif row["story_cell"]["merged_cell_propagation_applied"]:
            mapping_method = "merged_cell_propagation"
        else:
            matched_story_records = [
                stories_by_key[story_id.casefold()][0] for story_id in story_ids
            ]
            normalized_match = (
                row["story_cell"]["normalization_applied"]
                or row["story_cell"]["raw_value_type"] not in {"str", "NoneType"}
                or any(
                    record["normalization_applied"]
                    or record["raw_value_type"] != "str"
                    for record in matched_story_records
                )
            )
            mapping_method = "normalized_exact_id" if normalized_match else "exact_id"

        gold_candidates = (
            gold_by_stem.get(row["cq_id"].casefold(), []) if row["cq_id"] else []
        )
        if len(gold_candidates) == 1:
            gold = gold_candidates[0]
            gold_status = "mapped"
            gold_mapping_method = "exact_id"
            gold_path = str(gold["output_path"])
            gold_sha = gold["source_sha256"]
            gold_parse_success: bool | str = gold["parse_success"]
        elif len(gold_candidates) > 1:
            gold_status = "ambiguous"
            gold_mapping_method = "ambiguous"
            gold_path = "|".join(
                str(candidate["output_path"]) for candidate in gold_candidates
            )
            gold_sha = "|".join(
                candidate["source_sha256"] for candidate in gold_candidates
            )
            gold_parse_success = False
        else:
            gold_status = "missing"
            gold_mapping_method = "unmapped"
            gold_path = ""
            gold_sha = ""
            gold_parse_success = ""

        full_generation_included = not exclusion_reason
        gold_evaluable_included = (
            full_generation_included
            and gold_status == "mapped"
            and gold_parse_success is True
        )
        mapping_row = {
            "source_workbook": workbook_path.name,
            "source_sheet": "CQs",
            "source_row": row["source_row"],
            "raw_story_id": row["raw_story_id"],
            "story_id": row["story_id"],
            "cq_id": row["cq_id"],
            "cq_text_nonempty": str(bool(row["cq_text"])).lower(),
            "hidden_source_row": str(row["story_cell"]["hidden_row"]).lower(),
            "mapping_method": mapping_method,
            "merged_source_range": row["story_cell"]["merged_source_range"] or "",
            "full_generation_included": str(full_generation_included).lower(),
            "included_in_normalized_dataset": str(full_generation_included).lower(),
            "exclusion_reason": exclusion_reason,
            "exclusion_reasons": exclusion_reason,
            "missing_gold_only": str(
                full_generation_included and gold_status == "missing"
            ).lower(),
            "included_in_gold_evaluable_scope": str(gold_evaluable_included).lower(),
            "mapping_rule": "case-insensitive exact CQID-to-filename-stem",
            "gold_mapping_method": gold_mapping_method,
            "mapping_status": gold_status,
            "gold_path": gold_path,
            "gold_sha256": gold_sha,
            "gold_parse_success": gold_parse_success,
        }
        mapping_rows.append(mapping_row)

        story_records = [
            stories_by_key[story_id.casefold()][0]
            for story_id in story_ids
            if story_id.casefold() in stories_by_key
            and len(stories_by_key[story_id.casefold()]) == 1
        ]
        enriched = {
            **row,
            "story_ids": story_ids,
            "story_records": story_records,
            "mapping_method": mapping_method,
            "gold_path": gold_path or None,
            "gold_sha256": gold_sha or None,
            "gold_parse_success": gold_parse_success if gold_path else None,
            "gold_mapping_status": gold_status,
            "gold_evaluable_included": gold_evaluable_included,
        }
        if exclusion_reason:
            excluded_entry = {
                "source_sheet": "CQs",
                "source_row": row["source_row"],
                "story_id": row["story_id"],
                "cq_id": row["cq_id"],
                "exclusion_reason": exclusion_reason,
                "mapping_method": mapping_method,
                "hidden_source_row": row["story_cell"]["hidden_row"],
            }
            excluded.append(excluded_entry)
            errors.append(
                {
                    "category": exclusion_reason,
                    "source": f"{workbook_path.name}:CQs",
                    "identifier": str(row["source_row"]),
                    "detail": "Excluded from full generation scope by explicit source evidence.",
                }
            )
        else:
            full_generation_rows.append(enriched)

    groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for row in full_generation_rows:
        groups.setdefault(row["story_id"], []).append(row)

    normalized_items: list[dict[str, Any]] = []
    for story_id, rows in groups.items():
        scenario_stories: OrderedDict[str, dict[str, Any]] = OrderedDict()
        for row in rows:
            for story in row["story_records"]:
                scenario_stories.setdefault(story["story_id"], story)
        user_stories = [story["story_text"] for story in scenario_stories.values()]
        scenario = (
            user_stories[0]
            if len(scenario_stories) == 1
            else "\n\n".join(
                f"[{story['story_id']}]\n{story['story_text']}"
                for story in scenario_stories.values()
            )
        )
        cq_metadata = [
            {
                "cq_id": row["cq_id"],
                "cq_text": row["cq_text"],
                "category": row["category"] or None,
                "source_row": row["source_row"],
                "story_mapping_method": row["mapping_method"],
                "gold_module": row["gold_path"],
                "gold_sha256": row["gold_sha256"],
                "gold_parse_success": row["gold_parse_success"],
                "gold_mapping_status": row["gold_mapping_status"],
                "gold_evaluable": row["gold_evaluable_included"],
            }
            for row in rows
        ]
        normalized_items.append(
            {
                "dataset_id": dataset_id_for_story(story_id),
                "scenario_id": story_id,
                "scenario": scenario,
                "competency_questions": [row["cq_text"] for row in rows],
                "user_stories": user_stories,
                "constraints": {"output_format": "ttl"},
                "metadata": {
                    "dataset_scope": "full_generation",
                    "original_story_id": story_id,
                    "original_story_ids": list(scenario_stories),
                    "original_cq_ids": [row["cq_id"] for row in rows],
                    "source_workbook": workbook_path.name,
                    "source_sheet": "CQs",
                    "source_rows": [row["source_row"] for row in rows],
                    "story_source_rows": [
                        story["source_row"] for story in scenario_stories.values()
                    ],
                    "domain_identifier": None,
                    "domain_identifier_source": "not_provided_by_source_workbook",
                    "ontology_identifiers": [row["cq_id"] for row in rows],
                    "cq_records": cq_metadata,
                },
            }
        )

    mapped_gold_paths = {
        Path(row["gold_path"]).resolve()
        for row in mapping_rows
        if row["mapping_status"] == "mapped" and row["gold_path"]
    }
    orphan_modules = [
        {
            "archive": record["archive"],
            "archive_member": record["archive_member"],
            "output_path": str(record["output_path"]),
            "stem": record["stem"],
            "reason": (
                "No CQID with a normalized case-insensitive exact stem match"
                if record["stem"].casefold() not in all_source_cq_ids
                else "Matched CQ exists but mapping was not unique"
            ),
        }
        for record in gold_records
        if record["output_path"].resolve() not in mapped_gold_paths
    ]

    exclusion_counts = Counter(
        entry["exclusion_reason"] for entry in excluded
    )
    mapping_method_counts = Counter(row["mapping_method"] for row in mapping_rows)
    structure_summary = {
        key: value
        for key, value in structure.items()
        if key not in {"cq_story_cells", "story_id_cells"}
    }
    identifier_cells = [
        *structure["cq_story_cells"].values(),
        *structure["story_id_cells"].values(),
    ]
    string_identifier_values = [
        metadata["raw_value"]
        for metadata in identifier_cells
        if isinstance(metadata["raw_value"], str)
    ]
    case_only_match_count = 0
    for row in full_generation_rows:
        for story_id in row["story_ids"]:
            matched = stories_by_key[story_id.casefold()][0]["story_id"]
            if story_id != matched and story_id.casefold() == matched.casefold():
                case_only_match_count += 1
    normalization_inspection = {
        "normalization_policy": [
            "trim_surrounding_whitespace",
            "Unicode_NFC",
            "integer_numeric_to_canonical_text",
            "normalize_line_endings",
            "case_insensitive_exact_comparison",
        ],
        "story_id_cells_with_surrounding_whitespace": sum(
            value != value.strip() for value in string_identifier_values
        ),
        "story_id_cells_changed_by_unicode_nfc": sum(
            unicodedata.normalize("NFC", value) != value
            for value in string_identifier_values
        ),
        "story_id_cells_with_non_lf_line_endings": sum(
            "\r" in value for value in string_identifier_values
        ),
        "numeric_integer_story_id_cells": sum(
            metadata["raw_value_type"] == "int" for metadata in identifier_cells
        ),
        "integer_like_float_story_id_cells": sum(
            metadata["raw_value_type"] == "float"
            and isinstance(metadata["raw_value"], float)
            and metadata["raw_value"].is_integer()
            for metadata in identifier_cells
        ),
        "cq_id_representation_normalizations": sum(
            row["cq_id_normalization_applied"] for row in raw_cq_rows
        ),
        "case_only_story_matches": case_only_match_count,
        "fuzzy_or_semantic_matches": 0,
    }
    gold_evaluable_count = sum(
        row["included_in_gold_evaluable_scope"] == "true" for row in mapping_rows
    )
    audit = {
        "source_workbooks": [
            {"path": str(workbook_path), "sha256": sha256_file(workbook_path)}
        ],
        "sheets": [sheet_audit(sheet) for sheet in sheets],
        "workbook_structure_audit": structure_summary,
        "normalization_inspection": normalization_inspection,
        "source_row_count": max(0, cq_sheet.max_row - 1),
        "story_count": sum(len(records) for records in stories_by_key.values()),
        "cq_count": sum(1 for row in raw_cq_rows if row["cq_id"] and row["cq_text"]),
        "normalized_item_count": len(normalized_items),
        "normalized_cq_count": sum(
            len(item["competency_questions"]) for item in normalized_items
        ),
        "full_generation_scenario_count": len(normalized_items),
        "full_generation_cq_count": sum(
            len(item["competency_questions"]) for item in normalized_items
        ),
        "gold_evaluable_cq_count": gold_evaluable_count,
        "gold_module_count": len(gold_records),
        "exclusion_counts_by_reason": {
            category: exclusion_counts.get(category, 0)
            for category in exclusion_categories
        },
        "mapping_method_counts": {
            method: mapping_method_counts.get(method, 0)
            for method in [
                "exact_id",
                "normalized_exact_id",
                "merged_cell_propagation",
                "unmapped",
            ]
        },
        "missing_gold_count": sum(
            bool(row["cq_id"] and row["mapping_status"] == "missing")
            for row in mapping_rows
        ),
        "missing_story_count": sum(
            bool(row["cq_id"] and row["cq_text"] and not row["story_id"])
            for row in raw_cq_rows
        ),
        "unresolved_rows": excluded,
        "missing_gold_excludes_full_generation_count": sum(
            row["missing_gold_only"] == "true"
            and row["full_generation_included"] != "true"
            for row in mapping_rows
        ),
        "duplicate_ids": {
            "cq_ids": duplicate_cq_ids,
            "story_ids": duplicate_story_ids,
        },
        "missing_ids": {
            "cq_id_source_rows": [row["source_row"] for row in raw_cq_rows if not row["cq_id"]],
            "story_id_source_rows": [row["source_row"] for row in raw_cq_rows if not row["story_id"]],
        },
        "empty_cqs": [row["source_row"] for row in raw_cq_rows if not row["cq_text"]],
        "missing_stories": [
            {
                "source_row": row["source_row"],
                "cq_id": row["cq_id"],
                "story_id": row["story_id"],
            }
            for row in raw_cq_rows
            if row["cq_id"] and row["cq_text"] and not row["story_id"]
        ],
        "blank_story_rows": blank_story_rows,
        "missing_mappings": [
            {
                "source_row": row["source_row"],
                "story_id": row["story_id"],
                "cq_id": row["cq_id"],
            }
            for row in mapping_rows
            if row["cq_id"] and row["mapping_status"] == "missing"
        ],
        "orphan_modules": orphan_modules,
        "excluded_rows_and_reasons": excluded,
        "domain_identifier_column_present": False,
        "full_generation_scope": (
            "Every non-empty, non-duplicate CQ associated with story text by "
            "exact, normalized-exact, or merged-cell source evidence. Gold is optional."
        ),
        "gold_evaluable_scope": (
            "Rows in full generation scope with one parseable exact CQID-to-gold-stem match; "
            "gold_mapping.csv is the authoritative subset representation."
        ),
        "gold_mapping_rule": (
            "Dataset README states module filename is CQID.ttl; matching uses "
            "normalized, case-insensitive exact filename stem equality."
        ),
    }
    return normalized_items, mapping_rows, audit, errors


def logicalize_paths(
    value: Any,
    staging_root: Path,
    output_root: Path,
    repo_root: Path,
) -> Any:
    if isinstance(value, dict):
        return {
            key: logicalize_paths(item, staging_root, output_root, repo_root)
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [
            logicalize_paths(item, staging_root, output_root, repo_root)
            for item in value
        ]
    if isinstance(value, Path):
        return logicalize_paths(str(value), staging_root, output_root, repo_root)
    if isinstance(value, str):
        staging_text = str(staging_root)
        if value.startswith(staging_text):
            suffix = Path(value).relative_to(staging_root)
            return portable_path(output_root / suffix, repo_root)
        candidate = Path(value)
        if candidate.is_absolute():
            return portable_path(candidate, repo_root)
    return value


def enumerate_output_hashes(root: Path) -> list[dict[str, Any]]:
    records = []
    for path in sorted(candidate for candidate in root.rglob("*") if candidate.is_file()):
        records.append(
            {
                "path": str(path.relative_to(root)).replace("\\", "/"),
                "sha256": sha256_file(path),
                "bytes": path.stat().st_size,
            }
        )
    return records


def replace_managed_outputs(staging_root: Path, output_root: Path) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    managed = [
        "normalized",
        "gold",
        "prompts",
        "odps",
        "raw",
        "dataset_audit.json",
        "gold_mapping.csv",
        "source_row_reconciliation.csv",
        "conversion_errors.csv",
        "resource_manifest.json",
        "odp_manifest.json",
    ]
    resolved_root = output_root.resolve()
    for name in managed:
        source = staging_root / name
        destination = output_root / name
        if destination.exists():
            resolved_destination = destination.resolve()
            if resolved_root not in resolved_destination.parents:
                raise RuntimeError(f"Refusing to replace path outside output root: {destination}")
            if destination.is_dir():
                shutil.rmtree(destination)
            else:
                destination.unlink()
        if source.exists():
            os.replace(source, destination)


def prepare_dataset(
    repo_root: Path,
    dataset_source: Path,
    output_root: Path,
    approved_methods: Sequence[str] = DEFAULT_APPROVED_METHODS,
    password: str = DEFAULT_ARCHIVE_PASSWORD,
    enable_odps: bool = False,
) -> dict[str, Any]:
    repo_root = repo_root.resolve()
    dataset_source = dataset_source.resolve()
    output_root = output_root.resolve()
    workbook_files = sorted(dataset_source.glob("*.xlsx"))
    if len(workbook_files) != 1:
        raise ValueError(
            f"Expected exactly one XLSX in {dataset_source}; found {len(workbook_files)}"
        )
    workbook_path = workbook_files[0]

    staging_parent = output_root.parent
    staging_parent.mkdir(parents=True, exist_ok=True)
    staging_root = Path(
        tempfile.mkdtemp(prefix=f".{output_root.name}_prepare_", dir=staging_parent)
    )
    conversion_errors: list[dict[str, Any]] = []
    try:
        for directory in ("normalized", "gold", "prompts", "odps", "raw"):
            (staging_root / directory).mkdir(parents=True, exist_ok=True)

        gold_records, gold_errors = extract_gold_archives(
            dataset_source, staging_root / "gold", password
        )
        conversion_errors.extend(gold_errors)
        normalized_items, mapping_rows, audit, dataset_errors = (
            build_dataset_records_phase2b(workbook_path, gold_records)
        )
        conversion_errors.extend(dataset_errors)

        prompt_records, prompt_errors = prepare_prompts(
            repo_root, staging_root, approved_methods
        )
        conversion_errors.extend(prompt_errors)

        patterns_path = (
            repo_root / "external_resources" / "Ontogenia" / "data" / "patterns.csv"
        )
        procedure_path = (
            repo_root / "external_resources" / "Ontogenia" / "data" / "procedure.txt"
        )
        if enable_odps:
            odp_records, odp_errors = convert_patterns_to_odps(
                patterns_path, staging_root / "odps"
            )
            conversion_errors.extend(odp_errors)
            odp_reason = None
            applicable_methods = ["ontogenia-mp"]
        else:
            odp_records = []
            odp_reason = (
                "No approved Phase 2 course method calls _load_odps_text; "
                "ontogenia-mp is explicitly excluded from the approved three."
            )
            applicable_methods = []
            atomic_write_text(
                staging_root / "odps" / "README.md",
                "# ODP preparation\n\n"
                + odp_reason
                + "\n",
            )

        logical_normalized_items = logicalize_paths(
            normalized_items, staging_root, output_root, repo_root
        )
        normalized_payload = "".join(
            json.dumps(item, ensure_ascii=False, sort_keys=True) + "\n"
            for item in logical_normalized_items
        )
        atomic_write_text(
            staging_root / "normalized" / "project2_full_generation.jsonl",
            normalized_payload,
        )
        atomic_write_text(
            staging_root / "normalized" / "project2.jsonl",
            normalized_payload,
        )

        mapping_fields = [
            "source_workbook",
            "source_sheet",
            "source_row",
            "raw_story_id",
            "story_id",
            "cq_id",
            "cq_text_nonempty",
            "hidden_source_row",
            "mapping_method",
            "merged_source_range",
            "full_generation_included",
            "included_in_normalized_dataset",
            "exclusion_reason",
            "missing_gold_only",
            "included_in_gold_evaluable_scope",
            "mapping_rule",
            "gold_mapping_method",
            "mapping_status",
            "gold_path",
            "gold_sha256",
            "gold_parse_success",
            "exclusion_reasons",
        ]
        logical_mapping_rows = logicalize_paths(
            mapping_rows, staging_root, output_root, repo_root
        )
        write_csv(staging_root / "gold_mapping.csv", mapping_fields, logical_mapping_rows)
        reconciliation_fields = [
            "source_workbook",
            "source_sheet",
            "source_row",
            "raw_story_id",
            "story_id",
            "cq_id",
            "cq_text_nonempty",
            "hidden_source_row",
            "mapping_method",
            "merged_source_range",
            "full_generation_included",
            "exclusion_reason",
            "missing_gold_only",
            "included_in_gold_evaluable_scope",
            "gold_mapping_method",
            "mapping_status",
        ]
        write_csv(
            staging_root / "source_row_reconciliation.csv",
            reconciliation_fields,
            logical_mapping_rows,
        )

        for mapping_row in mapping_rows:
            if mapping_row["cq_id"] and mapping_row["mapping_status"] == "missing":
                conversion_errors.append(
                    {
                        "category": "missing_gold_mapping",
                        "source": f"{mapping_row['source_workbook']}:{mapping_row['source_sheet']}",
                        "identifier": mapping_row["cq_id"],
                        "detail": (
                            "No case-insensitive exact CQID-to-filename-stem match; "
                            "no identifier-prefix inference was attempted."
                        ),
                    }
                )

        error_fields = ["category", "source", "identifier", "detail"]
        logical_errors = logicalize_paths(
            conversion_errors, staging_root, output_root, repo_root
        )
        write_csv(
            staging_root / "conversion_errors.csv",
            error_fields,
            logical_errors,
        )

        logical_gold = logicalize_paths(
            gold_records, staging_root, output_root, repo_root
        )
        logical_odps = logicalize_paths(
            odp_records, staging_root, output_root, repo_root
        )
        logical_prompts = logicalize_paths(
            prompt_records, staging_root, output_root, repo_root
        )
        odp_manifest = {
            "schema_version": 1,
            "applicable_methods": applicable_methods,
            "not_applicable_reason": odp_reason,
            "source_patterns_csv": portable_path(patterns_path, repo_root),
            "source_patterns_sha256": sha256_file(patterns_path),
            "lookup_contract": {
                "directory": "datasets/ontology_generation/raw/ontogenia/odps",
                "candidate_order": ["name", "name.ttl", "name.owl", "name.rdf"],
            },
            "entries": logical_odps,
        }
        write_json(staging_root / "odp_manifest.json", odp_manifest)

        audit["missing_prompts"] = [
            record["method"] for record in prompt_records if record["status"] == "missing"
        ]
        audit["unparseable_gold_files"] = [
            {
                "archive": record["archive"],
                "archive_member": record["archive_member"],
                "error": record["parse_error"],
            }
            for record in gold_records
            if not record["parse_success"]
        ]
        audit["unparseable_odp_files"] = [
            {
                "original_name": record["original_name"],
                "filename": record["sanitized_filename"],
                "error": record["parse_error"],
            }
            for record in odp_records
            if not record["parse_success"]
        ]
        audit["approved_methods"] = list(approved_methods)
        audit["prompt_resources"] = logical_prompts
        audit["odp_applicable"] = enable_odps
        audit["conversion_error_count"] = len(conversion_errors)

        source_paths = [
            workbook_path,
            *sorted(dataset_source.glob("*.zip")),
            repo_root
            / "external_resources"
            / "Onto-Generation"
            / "PromptingTechniques"
            / "README.md",
            patterns_path,
            procedure_path,
        ]
        if "domain-ontogen" in approved_methods:
            source_paths.append(
                repo_root / "external_resources" / "Domain-OntoGen" / "README.md"
            )
        if "neon-gpt" in approved_methods:
            source_paths.extend(
                [
                    repo_root
                    / "external_resources"
                    / "NEON-GPT"
                    / "gpt_wine_ont_day1"
                    / "day1_gpt_prompt_list.txt",
                    repo_root / "external_resources" / "NEON-GPT" / "LICENSE",
                ]
            )
        source_files = []
        for path in source_paths:
            source_files.append(
                {
                    "path": portable_path(path, repo_root),
                    "sha256": sha256_file(path),
                    "bytes": path.stat().st_size,
                }
            )

        pre_manifest_outputs = enumerate_output_hashes(staging_root)
        manifest_basis = {
            "approved_methods": list(approved_methods),
            "authoritative_prompt_repositories": AUTHORITATIVE_PROMPT_REPOSITORIES,
            "source_files": source_files,
            "outputs": pre_manifest_outputs,
            "normalized_item_count": len(normalized_items),
            "gold_module_count": len(gold_records),
        }
        data_manifest_hash = canonical_json_hash(manifest_basis)
        audit["data_manifest_hash"] = data_manifest_hash
        logical_audit = logicalize_paths(
            audit, staging_root, output_root, repo_root
        )
        write_json(staging_root / "dataset_audit.json", logical_audit)

        resource_manifest = {
            "schema_version": 1,
            "data_manifest_hash": data_manifest_hash,
            "approved_methods": list(approved_methods),
            "authoritative_prompt_repositories": {
                method: AUTHORITATIVE_PROMPT_REPOSITORIES[method]
                for method in approved_methods
                if method in AUTHORITATIVE_PROMPT_REPOSITORIES
            },
            "source_files": source_files,
            "prompt_resources": logical_prompts,
            "gold_modules": logical_gold,
            "odp_manifest": "datasets/ontology_generation/odp_manifest.json",
            "outputs": enumerate_output_hashes(staging_root),
        }
        write_json(staging_root / "resource_manifest.json", resource_manifest)

        replace_managed_outputs(staging_root, output_root)
        return logical_audit
    finally:
        if staging_root.exists():
            shutil.rmtree(staging_root)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument(
        "--dataset-source",
        type=Path,
        default=Path("external_resources/Onto-Generation/Dataset_OntoGen"),
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("datasets/ontology_generation"),
    )
    parser.add_argument(
        "--approved-methods",
        default=",".join(DEFAULT_APPROVED_METHODS),
        help="Comma-separated implementation IDs frozen by the owner decision.",
    )
    parser.add_argument(
        "--archive-password",
        default=DEFAULT_ARCHIVE_PASSWORD,
    )
    parser.add_argument(
        "--enable-ontogenia-mp-odps",
        action="store_true",
        help="Prepare patterns.csv as ODP files only if ontogenia-mp is later approved.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = args.repo_root.resolve()
    dataset_source = (
        args.dataset_source
        if args.dataset_source.is_absolute()
        else repo_root / args.dataset_source
    )
    output_root = (
        args.output_root if args.output_root.is_absolute() else repo_root / args.output_root
    )
    approved_methods = tuple(
        method.strip() for method in args.approved_methods.split(",") if method.strip()
    )
    audit = prepare_dataset(
        repo_root=repo_root,
        dataset_source=dataset_source,
        output_root=output_root,
        approved_methods=approved_methods,
        password=args.archive_password,
        enable_odps=args.enable_ontogenia_mp_odps,
    )
    print(
        json.dumps(
            {
                "status": "prepared_with_reported_issues"
                if audit["conversion_error_count"]
                else "prepared",
                "full_generation_scenario_count": audit[
                    "full_generation_scenario_count"
                ],
                "full_generation_cq_count": audit["full_generation_cq_count"],
                "gold_evaluable_cq_count": audit["gold_evaluable_cq_count"],
                "gold_module_count": audit["gold_module_count"],
                "missing_prompt_count": len(audit["missing_prompts"]),
                "conversion_error_count": audit["conversion_error_count"],
                "data_manifest_hash": audit["data_manifest_hash"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
