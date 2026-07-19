from __future__ import annotations

import csv
import json
import sys
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

import pytest
from openpyxl import Workbook

from scripts.prepare_ontology_dataset import (
    build_dataset_records_phase2b,
    convert_patterns_to_odps,
    extract_domain_ontogen_prompt,
    inspect_workbook_with_openpyxl,
    prepare_dataset,
    prepare_prompts,
    read_xlsx,
    sha256_file,
)


REPO_ROOT = Path(__file__).resolve().parents[1]


def _column_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _sheet_xml(rows: list[list[str]]) -> str:
    max_columns = max(len(row) for row in rows)
    body = []
    for row_number, row in enumerate(rows, start=1):
        cells = []
        for column_number, value in enumerate(row, start=1):
            reference = f"{_column_name(column_number)}{row_number}"
            cells.append(
                f'<c r="{reference}" t="inlineStr"><is><t>{escape(value)}</t></is></c>'
            )
        body.append(f'<row r="{row_number}">{"".join(cells)}</row>')
    dimension = f"A1:{_column_name(max_columns)}{len(rows)}"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="{dimension}"/><sheetData>{"".join(body)}</sheetData>'
        "</worksheet>"
    )


def _write_workbook(path: Path) -> None:
    cqs = [
        ["StoryID", "CQID", "CQText", "Category of CQ"],
        ["S1", "CQ1", "Which actors participate?", "who"],
        ["", "", "", ""],
        ["", "PREFIX_2", "Which item lacks an explicit story?", "what"],
    ]
    stories = [
        ["StoryID", "StoryText"],
        ["S1", "A researcher documents participants."],
    ]
    with zipfile.ZipFile(path, "w") as workbook:
        workbook.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            "</Types>",
        )
        workbook.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="CQs" sheetId="1" r:id="rId1"/>'
            '<sheet name="Story" sheetId="2" r:id="rId2"/></sheets></workbook>',
        )
        workbook.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>'
            "</Relationships>",
        )
        workbook.writestr("xl/worksheets/sheet1.xml", _sheet_xml(cqs))
        workbook.writestr("xl/worksheets/sheet2.xml", _sheet_xml(stories))


def _write_fixture_repository(root: Path) -> tuple[Path, Path]:
    dataset_source = root / "external_resources/Onto-Generation/Dataset_OntoGen"
    dataset_source.mkdir(parents=True)
    _write_workbook(dataset_source / "Dataset.xlsx")
    turtle = b"@prefix ex: <https://example.org/> . ex:a ex:relatedTo ex:b .\n"
    with zipfile.ZipFile(dataset_source / "modules.zip", "w") as archive:
        archive.writestr("CQ1.ttl", turtle)
        archive.writestr("PREFIX.ttl", turtle)

    prompt_dir = root / "external_resources/Onto-Generation/PromptingTechniques"
    prompt_dir.mkdir(parents=True)
    (prompt_dir / "README.md").write_text(
        "## Memoryless CQbyCQ\n\n```python\n"
        '"Story: {story}\\nCQ: {CQ}\\nExisting RDF: {rdf}"\n'
        "```\n\n## Ontogenia\n",
        encoding="utf-8",
    )
    data_dir = root / "external_resources/Ontogenia/data"
    data_dir.mkdir(parents=True)
    (data_dir / "patterns.csv").write_text(
        "Name,Information,Pattern_owl\n", encoding="utf-8"
    )
    (data_dir / "procedure.txt").write_text("procedure\n", encoding="utf-8")
    return dataset_source, root / "datasets/ontology_generation"


def test_end_to_end_preserves_explicit_relationships_and_loader_contract(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    dataset_source, output_root = _write_fixture_repository(tmp_path)

    audit = prepare_dataset(
        repo_root=tmp_path,
        dataset_source=dataset_source,
        output_root=output_root,
        approved_methods=("ontogenia",),
    )

    assert audit["normalized_item_count"] == 1
    assert audit["normalized_cq_count"] == 1
    assert audit["orphan_modules"][0]["stem"] == "PREFIX"
    assert {entry["source_row"] for entry in audit["excluded_rows_and_reasons"]} == {
        3,
        4,
    }
    excluded_by_row = {
        entry["source_row"]: entry for entry in audit["excluded_rows_and_reasons"]
    }
    assert excluded_by_row[4]["exclusion_reason"] == "missing_story_id"

    dataset_path = output_root / "normalized/project2.jsonl"
    item = json.loads(dataset_path.read_text(encoding="utf-8").strip())
    assert item["scenario_id"] == "S1"
    assert item["metadata"]["original_cq_ids"] == ["CQ1"]
    assert item["metadata"]["cq_records"][0]["gold_module"].startswith(
        "datasets/ontology_generation/gold/"
    )
    assert str(tmp_path).replace("\\", "/") not in json.dumps(item)

    monkeypatch.syspath_prepend(str(REPO_ROOT / "restapi"))
    from app.utils.ontology_dataset import load_ontology_items

    loaded = load_ontology_items(str(dataset_path))
    assert len(loaded) == 1
    assert loaded[0].competency_questions == ["Which actors participate?"]

    with (output_root / "gold_mapping.csv").open(encoding="utf-8", newline="") as handle:
        mapping = list(csv.DictReader(handle))
    missing_story_row = next(row for row in mapping if row["cq_id"] == "PREFIX_2")
    assert missing_story_row["mapping_status"] == "missing"
    assert missing_story_row["included_in_normalized_dataset"] == "false"


def test_workbook_reader_preserves_physical_blank_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "Dataset.xlsx"
    _write_workbook(workbook)
    sheets = {sheet.name: sheet for sheet in read_xlsx(workbook)}
    assert sheets["CQs"].max_row == 4
    assert sheets["CQs"].rows[3] == ["", "", "", ""]


def test_odp_conversion_sanitizes_names_and_reports_invalid_rdf(tmp_path: Path) -> None:
    patterns = tmp_path / "patterns.csv"
    with patterns.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Name", "Information", "Pattern_owl"])
        writer.writeheader()
        writer.writerow(
            {
                "Name": "Actor / Role",
                "Information": "valid",
                "Pattern_owl": (
                    '<?xml version="1.0"?><rdf:RDF '
                    'xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"/>'
                ),
            }
        )
        writer.writerow(
            {"Name": "Broken", "Information": "invalid", "Pattern_owl": "not rdf"}
        )

    records, errors = convert_patterns_to_odps(patterns, tmp_path / "odps")

    assert records[0]["sanitized_filename"] == "Actor_Role.owl"
    assert records[0]["parse_success"] is True
    assert records[1]["parse_success"] is False
    assert errors[0]["category"] == "unparseable_odp_file"


def test_prepared_project2_dataset_loads_and_matches_its_audit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_root = REPO_ROOT / "datasets/ontology_generation"
    dataset_path = output_root / "normalized/project2_full_generation.jsonl"
    audit = json.loads((output_root / "dataset_audit.json").read_text(encoding="utf-8"))

    monkeypatch.syspath_prepend(str(REPO_ROOT / "restapi"))
    from app.utils.ontology_dataset import load_ontology_items

    loaded = load_ontology_items(str(dataset_path))
    assert len(loaded) == audit["normalized_item_count"]
    assert sum(len(item.competency_questions) for item in loaded) == audit[
        "normalized_cq_count"
    ]
    assert not audit["unparseable_gold_files"]
    assert audit["missing_prompts"] == []
    assert audit["full_generation_scenario_count"] == 17
    assert audit["full_generation_cq_count"] == 74
    assert audit["gold_evaluable_cq_count"] == 27
    assert audit["exclusion_counts_by_reason"]["missing_story_id"] == 38
    assert audit["exclusion_counts_by_reason"]["empty_cq"] == 6
    assert audit["exclusion_counts_by_reason"] == {
        "ambiguous_story_mapping": 0,
        "duplicate_cq": 0,
        "empty_cq": 6,
        "invalid_source_row": 0,
        "missing_gold_only": 0,
        "missing_story_id": 38,
        "other": 0,
        "story_id_not_found": 0,
    }
    assert audit["mapping_method_counts"] == {
        "exact_id": 58,
        "normalized_exact_id": 16,
        "merged_cell_propagation": 0,
        "unmapped": 44,
    }
    assert audit["missing_gold_excludes_full_generation_count"] == 0
    assert audit["workbook_structure_audit"]["openpyxl_inspection_performed"] is True
    assert audit["workbook_structure_audit"]["sheets"]["CQs"][
        "hidden_row_ranges"
    ] == ["2-17", "29-50", "53-58"]
    assert audit["workbook_structure_audit"]["sheets"]["CQs"][
        "story_id_column_merged_ranges"
    ] == []
    assert audit["workbook_structure_audit"]["sheets"]["CQs"][
        "formula_cells"
    ] == []
    assert audit["normalization_inspection"]["numeric_integer_story_id_cells"] == 32
    assert audit["normalization_inspection"]["story_id_cells_with_surrounding_whitespace"] == 0
    assert audit["normalization_inspection"]["story_id_cells_changed_by_unicode_nfc"] == 0
    assert audit["normalization_inspection"]["case_only_story_matches"] == 0
    assert audit["normalization_inspection"]["fuzzy_or_semantic_matches"] == 0

    serialized_outputs = "\n".join(
        path.read_text(encoding="utf-8", errors="replace")
        for path in output_root.rglob("*")
        if path.is_file()
    )
    assert "A:/Projects_File" not in serialized_outputs
    assert "C:/Users/" not in serialized_outputs


def test_authoritative_prompt_recovery_is_exact_and_deterministic(
    tmp_path: Path,
) -> None:
    domain_source = REPO_ROOT / "external_resources/Domain-OntoGen/README.md"
    domain_prompt = extract_domain_ontogen_prompt(domain_source)
    assert domain_prompt
    assert domain_prompt.count("{CQ}") == 1
    assert domain_prompt.count("{OS}") == 1
    assert domain_prompt.endswith("O: \n")

    first_root = tmp_path / "first"
    second_root = tmp_path / "second"
    first_records, first_errors = prepare_prompts(
        REPO_ROOT, first_root, ("domain-ontogen", "neon-gpt")
    )
    second_records, second_errors = prepare_prompts(
        REPO_ROOT, second_root, ("domain-ontogen", "neon-gpt")
    )
    assert not first_errors
    assert not second_errors
    assert first_records == second_records

    domain_copy = first_root / "raw/domain-ontogen/prompt.txt"
    neon_source = (
        REPO_ROOT
        / "external_resources/NEON-GPT/gpt_wine_ont_day1/day1_gpt_prompt_list.txt"
    )
    neon_copy = first_root / "raw/neon-gpt/day1_gpt_prompt_list.txt"
    assert sha256_file(domain_copy) == (
        "f9e3945421508cd6a82613caf0d26fe802084178d950b2f1bd81b0446c2add4e"
    )
    assert neon_copy.read_bytes() == neon_source.read_bytes()
    assert sha256_file(neon_copy) == (
        "40d0baf11f4945fc37f0a4d2f67a7efbbf3a249e0ae8e5b105672ee79a83f44a"
    )


def test_merged_story_id_propagation_is_explicit_source_evidence(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "Dataset.xlsx"
    workbook = Workbook()
    cqs = workbook.active
    cqs.title = "CQs"
    cqs.append(["StoryID", "CQID", "CQText", "Category of CQ"])
    cqs.append(["S1", "CQ1", "Question one?", "what"])
    cqs.append([None, "CQ2", "Question two?", "what"])
    cqs.merge_cells("A2:A3")
    stories = workbook.create_sheet("Story")
    stories.append(["StoryID", "StoryText"])
    stories.append(["S1", "Explicit merged-cell scenario."])
    workbook.save(workbook_path)

    structure = inspect_workbook_with_openpyxl(workbook_path, 3, 2)
    assert structure["cq_story_cells"][3]["normalized_value"] == "S1"
    assert structure["cq_story_cells"][3]["merged_cell_propagation_applied"] is True

    items, mapping, audit, errors = build_dataset_records_phase2b(workbook_path, [])
    assert not errors
    assert len(items) == 1
    assert audit["full_generation_cq_count"] == 2
    assert audit["mapping_method_counts"] == {
        "exact_id": 1,
        "normalized_exact_id": 0,
        "merged_cell_propagation": 1,
        "unmapped": 0,
    }
    assert mapping[1]["mapping_method"] == "merged_cell_propagation"
